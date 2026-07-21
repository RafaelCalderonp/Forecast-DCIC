"""
Fase 1 — Script de carga inicial
1. Aplica migration_fase1.sql
2. Crea hash bcrypt para usuario admin
3. Borra TODOS los productos actuales (y sus dependencias)
4. Carga 716 productos + forecast 2026 desde el Excel

Uso:
    python cargar_excel_fase1.py --excel "ruta/al/Forecast_2026 v16.xlsx"
    python cargar_excel_fase1.py --excel "ruta/al/Forecast_2026 v16.xlsx" --solo-productos
    python cargar_excel_fase1.py --excel "ruta/al/Forecast_2026 v16.xlsx" --solo-forecast
"""

import sys
import os
import re
import asyncio
import argparse
import getpass
import openpyxl
from decimal import Decimal, InvalidOperation
from pathlib import Path

import bcrypt
import asyncpg

# ── Configuración ───────────────────────────────────────────────────────────
from dotenv import load_dotenv
load_dotenv()
DATABASE_URL = os.getenv(
    "DATABASE_URL",
    "postgresql://postgres:postgres@localhost:5432/forecast_dcic"
).replace("postgresql+asyncpg://", "postgresql://")

EXCEL_DEFAULT = r"C:\Users\rafae\OneDrive - IMPORTADORA DCIC SPA\DCIC SpA\001. Analisis Informe\Forecast_2026 v16.xlsx"

MIGRATION_SQL = Path(__file__).parent.parent / "database" / "migration_fase1.sql"

TEMPORADAS_MAP = {
    "verano":           "Verano",
    "invierno":         "Invierno",
    "no estacional":    "No Estacional",
    "no-estacional":    "No Estacional",
    "noestacional":     "No Estacional",
    "verano/rotativo":  "Verano/Rotativo",
    "rotativo":         "Verano/Rotativo",
    "verano rotativo":  "Verano/Rotativo",
}

MESES = ["ene","feb","mar","abr","may","jun","jul","ago","sep","oct","nov","dic"]


# ── Helpers ──────────────────────────────────────────────────────────────────

def limpiar(valor):
    if valor is None:
        return None
    return str(valor).strip()

MAX_NUMERIC = Decimal("9999999999.99")  # max para NUMERIC(12,2)

def a_decimal(valor):
    if valor is None or str(valor).strip() == "":
        return None
    try:
        txt = str(valor).replace("$","").replace(".","").replace(",",".").strip()
        d = Decimal(txt)
        if abs(d) > MAX_NUMERIC:
            return None  # desbordamiento — ignorar
        return d
    except InvalidOperation:
        return None

def a_int(valor):
    if valor is None or str(valor).strip() == "":
        return 0
    try:
        return int(float(str(valor).replace(",","").replace(".","").strip()))
    except (ValueError, TypeError):
        return 0

def a_bool(valor):
    if valor is None:
        return False
    v = str(valor).strip().lower()
    return v in ("si","sí","yes","true","1","x","✓","verdadero")

def normalizar_temporada(valor):
    if not valor:
        return None
    v = str(valor).strip().lower()
    for k, nombre in TEMPORADAS_MAP.items():
        if k in v:
            return nombre
    return None

def normalizar_pareto(valor):
    if not valor:
        return None
    v = str(valor).strip().upper()
    if v in ("A","B","C"):
        return v
    return None


# ── Leer Excel ────────────────────────────────────────────────────────────────

def leer_excel(ruta_excel):
    print(f"\n[EXCEL] Abriendo: {ruta_excel}")
    wb = openpyxl.load_workbook(ruta_excel, data_only=True)

    # Buscar hoja principal
    hoja_nombre = None
    for nombre in wb.sheetnames:
        if "nuevo" in nombre.lower() or nombre.strip() == "Forecast 2026":
            if "nuevo" in nombre.lower():
                hoja_nombre = nombre
                break
            hoja_nombre = nombre
    if not hoja_nombre:
        hoja_nombre = wb.sheetnames[0]
    print(f"   Hoja principal: '{hoja_nombre}'")

    ws = wb[hoja_nombre]

    # Detectar fila de cabeceras (busca "CODIGO" o "SKU" en col B)
    fila_header = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        for celda in row:
            if celda and str(celda).strip().upper() in ("CODIGO", "SKU", "CÓDIGO"):
                fila_header = i
                break
        if fila_header:
            break

    if not fila_header:
        raise ValueError("No se encontró fila de cabeceras con 'CODIGO' en las primeras 10 filas")

    print(f"   Fila de cabeceras: {fila_header}")

    # Leer cabeceras
    cabeceras_raw = [ws.cell(fila_header, c).value for c in range(1, ws.max_column + 1)]
    cabeceras = [str(h).strip() if h else "" for h in cabeceras_raw]

    # Mapear columnas fijas
    def col(nombre_buscar):
        nombre_buscar = nombre_buscar.lower()
        for i, h in enumerate(cabeceras):
            if nombre_buscar in h.lower():
                return i
        return None

    idx = {
        "marca":       col("marca"),
        "sku":         col("codigo"),
        "tipo":        col("tipo"),
        "temporada":   col("temporal"),
        "descripcion": col("descripcion"),
        "categoria":   col("categoria"),
        "subcategoria":col("sub categ"),
        "precio_lp":   col("precio lp"),
        "precio_prom": col("precio pro"),
        "precio_cyber":col("minimo cyber"),
        "con_piedras": col("piedras reliq"),
        "lp_liquid":   col("lp liquid"),
        "pareto":      col("pareto"),
        "prioridad":   col("prioridad"),
        "recomendacion":col("recomend"),
        "rehacer":     col("rehacer"),
        "cerrado":     col("cerrado"),
        "al_31_mayo":  col("31 mayo"),
        "piedras":     col("piedras") if col("piedras") != col("piedras reliq") else None,
        "descontinuado":col("descontinuado"),
        "comentario":  col("comentario"),
    }

    # Detectar columnas de forecast y ventas por mes
    # Patrón: "Forecast ene26", "vta ene24", "vta ene25", "Ajuste ene", "PxQ ene", "Consumo Pack"
    meses_cols = {}  # {mes_idx_0based: {tipo: col_idx}}
    for i, h in enumerate(cabeceras):
        hl = h.lower().strip()
        for mi, mes in enumerate(MESES):
            if f"forecast {mes}" in hl or f"forecast\n{mes}" in hl:
                meses_cols.setdefault(mi, {})["forecast"] = i
            elif f"ajuste {mes}" in hl or f"ajuste\n{mes}" in hl:
                meses_cols.setdefault(mi, {})["ajuste"] = i
            elif f"pxq {mes}" in hl or f"pxq\n{mes}" in hl:
                meses_cols.setdefault(mi, {})["pxq"] = i
            elif f"vta {mes}25" in hl or f"vta\n{mes}25" in hl:
                meses_cols.setdefault(mi, {})["vta25"] = i
            elif f"vta {mes}24" in hl or f"vta\n{mes}24" in hl:
                meses_cols.setdefault(mi, {})["vta24"] = i
            elif f"consumo" in hl and mes in hl:
                meses_cols.setdefault(mi, {})["consumo_pack"] = i

    print(f"   Meses detectados con forecast: {sorted(meses_cols.keys())}")

    # Leer filas de datos
    productos = []
    errores = []

    for row_num in range(fila_header + 1, ws.max_row + 1):
        fila = [ws.cell(row_num, c).value for c in range(1, ws.max_column + 1)]

        # SKU en columna B (idx["sku"])
        if idx["sku"] is None:
            continue
        sku_raw = limpiar(fila[idx["sku"]])
        if not sku_raw or sku_raw.upper().startswith("TOTAL"):
            continue

        # Validar formato SKU (R#### o R#####)
        if not re.match(r'^R\d{4,5}$', sku_raw, re.IGNORECASE):
            # Permitir igual pero marcar
            pass

        marca = limpiar(fila[idx["marca"]]) if idx["marca"] is not None else None
        if not marca:
            errores.append(f"Fila {row_num}: SKU '{sku_raw}' sin marca — omitido")
            continue

        temporada_raw = limpiar(fila[idx["temporada"]]) if idx["temporada"] is not None else None
        temporada_nombre = normalizar_temporada(temporada_raw)

        forecast_meses = {}
        for mi, cols in meses_cols.items():
            forecast_meses[mi] = {
                "forecast":      a_int(fila[cols["forecast"]]) if "forecast" in cols else 0,
                "ajuste":        a_int(fila[cols["ajuste"]])   if "ajuste"   in cols else 0,
                "vta25":         a_int(fila[cols["vta25"]])    if "vta25"    in cols else 0,
                "vta24":        a_int(fila[cols["vta24"]])     if "vta24"    in cols else 0,
                "consumo_pack":  a_int(fila[cols["consumo_pack"]]) if "consumo_pack" in cols else 0,
                "pxq":           a_decimal(fila[cols["pxq"]]) if "pxq" in cols else None,
            }

        productos.append({
            "sku":             sku_raw,
            "marca":           marca,
            "tipo":            limpiar(fila[idx["tipo"]]) if idx["tipo"] is not None else "Producto",
            "temporada":       temporada_nombre,
            "descripcion":     limpiar(fila[idx["descripcion"]]) if idx["descripcion"] is not None else None,
            "categoria":       limpiar(fila[idx["categoria"]]) if idx["categoria"] is not None else None,
            "subcategoria":    limpiar(fila[idx["subcategoria"]]) if idx["subcategoria"] is not None else None,
            "precio_lp":       a_decimal(fila[idx["precio_lp"]]) if idx["precio_lp"] is not None else None,
            "precio_prom":     a_decimal(fila[idx["precio_prom"]]) if idx["precio_prom"] is not None else None,
            "precio_cyber":    a_decimal(fila[idx["precio_cyber"]]) if idx["precio_cyber"] is not None else None,
            "precio_liquid":   a_decimal(fila[idx["lp_liquid"]]) if idx["lp_liquid"] is not None else None,
            "con_piedras":     a_bool(fila[idx["con_piedras"]]) if idx["con_piedras"] is not None else False,
            "pareto":          normalizar_pareto(fila[idx["pareto"]]) if idx["pareto"] is not None else None,
            "prioridad":       a_int(fila[idx["prioridad"]]) if idx["prioridad"] is not None else 0,
            "recomendacion":   limpiar(fila[idx["recomendacion"]]) if idx["recomendacion"] is not None else None,
            "rehacer":         a_bool(fila[idx["rehacer"]]) if idx["rehacer"] is not None else False,
            "cerrado":         a_bool(fila[idx["cerrado"]]) if idx["cerrado"] is not None else False,
            "al_31_mayo":      a_int(fila[idx["al_31_mayo"]]) if idx["al_31_mayo"] is not None else 0,
            "es_piedra":       a_bool(fila[idx["piedras"]]) if idx["piedras"] is not None else False,
            "activo":          not a_bool(fila[idx["descontinuado"]]) if idx["descontinuado"] is not None else True,
            "comentario":      limpiar(fila[idx["comentario"]]) if idx["comentario"] is not None else None,
            "forecast_meses":  forecast_meses,
        })

    print(f"   Productos leídos: {len(productos)}")
    if errores:
        print(f"   [WARN] Advertencias ({len(errores)}):")
        for e in errores[:10]:
            print(f"     {e}")
        if len(errores) > 10:
            print(f"     ... y {len(errores)-10} más")

    return productos


# ── Funciones de BD ───────────────────────────────────────────────────────────

async def get_or_create(conn, tabla, campo, valor):
    """Retorna el id de un registro, creándolo si no existe."""
    row = await conn.fetchrow(f"SELECT id FROM {tabla} WHERE {campo} = $1", valor)
    if row:
        return row["id"]
    row = await conn.fetchrow(
        f"INSERT INTO {tabla} ({campo}) VALUES ($1) ON CONFLICT ({campo}) DO UPDATE SET {campo}=EXCLUDED.{campo} RETURNING id",
        valor
    )
    return row["id"]

async def get_or_create_subcategoria(conn, nombre, categoria_id):
    row = await conn.fetchrow(
        "SELECT id FROM subcategorias WHERE nombre = $1 AND categoria_id = $2",
        nombre, categoria_id
    )
    if row:
        return row["id"]
    row = await conn.fetchrow(
        "INSERT INTO subcategorias (nombre, categoria_id) VALUES ($1, $2) "
        "ON CONFLICT (nombre, categoria_id) DO UPDATE SET nombre=EXCLUDED.nombre RETURNING id",
        nombre, categoria_id
    )
    return row["id"]


# ── Main async ────────────────────────────────────────────────────────────────

async def main(args):
    conn = await asyncpg.connect(DATABASE_URL)
    print(f"\n[OK] Conectado a la BD")

    try:
        # ── Paso 1: Aplicar migración ────────────────────────────────────
        if not args.solo_productos and not args.solo_forecast:
            print("\n[MIGRA] PASO 1: Aplicando migration_fase1.sql ...")
            sql_migracion = MIGRATION_SQL.read_text(encoding="utf-8")

            # Verificar si ya se aplicó
            check = await conn.fetchval(
                "SELECT EXISTS (SELECT 1 FROM information_schema.tables WHERE table_name = 'roles')"
            )
            if check:
                print("   [WARN] Migración ya aplicada (tabla 'roles' existe) — omitiendo")
            else:
                try:
                    await conn.execute(sql_migracion)
                    print("   [OK] Migración aplicada correctamente")
                except Exception as e:
                    print(f"   [ERR] Error en migración: {e}")
                    raise

        # ── Paso 2: Hash admin password ──────────────────────────────────
        if not args.solo_productos and not args.solo_forecast:
            print("\n[AUTH] PASO 2: Configurando usuario admin ...")
            pwd = args.admin_password or getpass.getpass("   Ingresa password para admin@dcic.cl: ")
            pwd_hash = bcrypt.hashpw(pwd.encode(), bcrypt.gensalt()).decode()
            await conn.execute(
                "UPDATE usuarios SET password_hash = $1 WHERE email = 'admin@dcic.cl'",
                pwd_hash
            )
            print("   [OK] Password admin configurado")

        # ── Paso 3: Borrar todos los productos actuales ──────────────────
        if not args.solo_forecast:
            print("\n[DEL]  PASO 3: Borrando productos actuales ...")
            counts = {}
            for tabla in ["plan_compras","forecast_canal","forecast","ventas","stock","pack_componentes","packs","productos"]:
                try:
                    r = await conn.execute(f"DELETE FROM {tabla}")
                    n = int(r.split()[-1])
                    if n > 0:
                        counts[tabla] = n
                except Exception:
                    pass
            for tabla, n in counts.items():
                print(f"   Borrados {n:>5} registros de '{tabla}'")
            print("   [OK] Limpieza completa")

            # También limpiar marcas, categorías, subcategorías vacías
            for tabla in ["subcategorias","categorias","marcas"]:
                try:
                    await conn.execute(f"DELETE FROM {tabla}")
                except Exception:
                    pass
            print("   [OK] Marcas y categorías limpiadas")

        # ── Paso 4: Leer Excel ───────────────────────────────────────────
        print("\n[DATA] PASO 4: Leyendo Excel ...")
        productos = leer_excel(args.excel)

        if not productos:
            print("[ERR] No se encontraron productos en el Excel. Verifica el archivo.")
            return

        # ── Paso 5: Cargar productos ─────────────────────────────────────
        if not args.solo_forecast:
            print(f"\n[PROD] PASO 5: Cargando {len(productos)} productos ...")
            temporada_cache = {}
            cargados = errores_carga = 0

            # Pre-cargar todos los lookups
            for p in productos:
                await get_or_create(conn, "marcas",     "nombre", p["marca"])
                await get_or_create(conn, "categorias", "nombre", p["categoria"] or "Sin Categoria")
                if p["subcategoria"]:
                    cat_id = await conn.fetchval("SELECT id FROM categorias WHERE nombre=$1", p["categoria"] or "Sin Categoria")
                    await get_or_create_subcategoria(conn, p["subcategoria"], cat_id)

            # Cargar temporadas cache
            rows_temp = await conn.fetch("SELECT id, nombre FROM temporadas")
            temporada_cache = {r["nombre"]: r["id"] for r in rows_temp}

            # Construir lista de tuplas para executemany
            filas_prod = []
            for p in productos:
                marca_id     = await conn.fetchval("SELECT id FROM marcas WHERE nombre=$1", p["marca"])
                categoria_id = await conn.fetchval("SELECT id FROM categorias WHERE nombre=$1", p["categoria"] or "Sin Categoria")
                subcategoria_id = None
                if p["subcategoria"]:
                    subcategoria_id = await conn.fetchval(
                        "SELECT id FROM subcategorias WHERE nombre=$1 AND categoria_id=$2",
                        p["subcategoria"], categoria_id
                    )
                temporada_id = temporada_cache.get(p["temporada"]) if p["temporada"] else None

                filas_prod.append((
                    p["sku"], marca_id, categoria_id, subcategoria_id, temporada_id,
                    p["tipo"] or "Producto", p["descripcion"],
                    p["precio_lp"] or Decimal(0), p["precio_prom"] or Decimal(0),
                    p["precio_cyber"], p["precio_liquid"],
                    p["con_piedras"],
                    p["pareto"], p["prioridad"], 0,
                    p["recomendacion"], p["comentario"],
                    p["es_piedra"], p["rehacer"], p["cerrado"],
                    p["al_31_mayo"], p["activo"]
                ))

            async with conn.transaction():
                await conn.executemany("""
                    INSERT INTO productos (
                        sku, marca_id, categoria_id, subcategoria_id, temporada_id,
                        tipo, descripcion,
                        precio_venta_bruto, precio_venta_neto,
                        precio_minimo_evento, precio_liquidacion,
                        con_piedras_reliquidacion,
                        grupo_pareto, prioridad_compra, moq,
                        recomendacion, comentario_compra,
                        es_piedra, rehacer_forecast, cerrado,
                        al_31_mayo, activo
                    ) VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16,$17,$18,$19,$20,$21,$22)
                    ON CONFLICT (sku) DO UPDATE SET
                        marca_id=EXCLUDED.marca_id, categoria_id=EXCLUDED.categoria_id,
                        subcategoria_id=EXCLUDED.subcategoria_id, temporada_id=EXCLUDED.temporada_id,
                        tipo=EXCLUDED.tipo, descripcion=EXCLUDED.descripcion,
                        precio_venta_bruto=EXCLUDED.precio_venta_bruto,
                        precio_venta_neto=EXCLUDED.precio_venta_neto,
                        precio_minimo_evento=EXCLUDED.precio_minimo_evento,
                        precio_liquidacion=EXCLUDED.precio_liquidacion,
                        con_piedras_reliquidacion=EXCLUDED.con_piedras_reliquidacion,
                        grupo_pareto=EXCLUDED.grupo_pareto,
                        prioridad_compra=EXCLUDED.prioridad_compra,
                        recomendacion=EXCLUDED.recomendacion,
                        comentario_compra=EXCLUDED.comentario_compra,
                        es_piedra=EXCLUDED.es_piedra, rehacer_forecast=EXCLUDED.rehacer_forecast,
                        cerrado=EXCLUDED.cerrado, al_31_mayo=EXCLUDED.al_31_mayo,
                        activo=EXCLUDED.activo, updated_at=NOW()
                """, filas_prod)
                cargados = len(filas_prod)

            print(f"   [OK] Productos cargados: {cargados} | Errores: {errores_carga}")

        # ── Paso 6: Cargar forecast ──────────────────────────────────────
        print(f"\n[FC] PASO 6: Cargando forecast 2026 ...")
        fc_cargados = fc_errores = 0

        filas_fc = []
        for p in productos:
            for mi, datos in p["forecast_meses"].items():
                filas_fc.append((
                    p["sku"], 2026, mi + 1,
                    datos["forecast"], datos["ajuste"],
                    datos["vta25"], datos["vta24"],
                    datos["consumo_pack"], datos["pxq"]
                ))

        async with conn.transaction():
            await conn.executemany("""
                INSERT INTO forecast (sku, anio, mes, cantidad, ajuste,
                    venta_anio_anterior, venta_dos_anios, consumo_pack, pxq)
                VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9)
                ON CONFLICT (sku, anio, mes) DO UPDATE SET
                    cantidad=EXCLUDED.cantidad, ajuste=EXCLUDED.ajuste,
                    venta_anio_anterior=EXCLUDED.venta_anio_anterior,
                    venta_dos_anios=EXCLUDED.venta_dos_anios,
                    consumo_pack=EXCLUDED.consumo_pack, pxq=EXCLUDED.pxq,
                    updated_at=NOW()
            """, filas_fc)
            fc_cargados = len(filas_fc)

        if False:  # bloque original eliminado
            for p in productos:
                for mi, datos in p["forecast_meses"].items():
                    try:
                        async with conn.transaction():
                            await conn.execute("""
                            INSERT INTO forecast (
                                sku, anio, mes,
                                cantidad, ajuste,
                                venta_anio_anterior, venta_dos_anios,
                                consumo_pack, pxq
                            ) VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9)
                            ON CONFLICT (sku, anio, mes) DO UPDATE SET
                                cantidad=EXCLUDED.cantidad,
                                ajuste=EXCLUDED.ajuste,
                                venta_anio_anterior=EXCLUDED.venta_anio_anterior,
                                venta_dos_anios=EXCLUDED.venta_dos_anios,
                                consumo_pack=EXCLUDED.consumo_pack,
                                pxq=EXCLUDED.pxq,
                                updated_at=NOW()
                        """,
                            p["sku"], 2026, mi + 1,
                            datos["forecast"], datos["ajuste"],
                            datos["vta25"], datos["vta24"],
                            datos["consumo_pack"], datos["pxq"]
                        )
                        pass
                    except Exception:
                        pass

        print(f"   [OK] Filas forecast cargadas: {fc_cargados} | Errores: {fc_errores}")

        # ── Verificación final ───────────────────────────────────────────
        print("\n[DATA] VERIFICACIÓN FINAL:")
        n_prod    = await conn.fetchval("SELECT COUNT(*) FROM productos")
        n_activos = await conn.fetchval("SELECT COUNT(*) FROM productos WHERE activo = TRUE")
        n_fc      = await conn.fetchval("SELECT COUNT(*) FROM forecast WHERE anio = 2026")
        n_marcas  = await conn.fetchval("SELECT COUNT(*) FROM marcas")
        n_cats    = await conn.fetchval("SELECT COUNT(*) FROM categorias")
        print(f"   Productos:   {n_prod:>5} total | {n_activos} activos")
        print(f"   Forecast:    {n_fc:>5} filas (esperado ~{len(productos)*12})")
        print(f"   Marcas:      {n_marcas:>5}")
        print(f"   Categorías:  {n_cats:>5}")
        print("\n[OK] FASE 1 COMPLETADA\n")

    finally:
        await conn.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Carga inicial Forecast DCIC — Fase 1")
    parser.add_argument("--excel", default=EXCEL_DEFAULT, help="Ruta al Excel")
    parser.add_argument("--admin-password", default=None, help="Password admin (si no se pasa, pide por consola)")
    parser.add_argument("--solo-productos", action="store_true", help="Solo cargar productos (omite migración)")
    parser.add_argument("--solo-forecast",  action="store_true", help="Solo cargar forecast (omite productos)")
    args = parser.parse_args()

    if not Path(args.excel).exists():
        print(f"[ERR] No se encontró el Excel: {args.excel}")
        sys.exit(1)

    asyncio.run(main(args))
