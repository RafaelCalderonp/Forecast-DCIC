# routers/lista_precios.py
from fastapi import APIRouter, Depends, UploadFile, File, Form, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text
from typing import Optional
from io import BytesIO
import openpyxl

from database import get_db
from auth import require_rol

router = APIRouter()

CAMPOS_PRECIO = {
    "precio_venta_bruto":   ["precio_venta_bruto", "precio bruto", "pvp bruto", "precio venta bruto"],
    "precio_venta_neto":    ["precio_venta_neto",  "precio neto",  "pvp neto",  "precio venta neto"],
    "costo_unitario_neto":  ["costo_unitario_neto","costo neto",   "costo",     "costo unitario"],
    "precio_minimo_evento": ["precio_minimo_evento","precio minimo","precio mínimo","precio evento"],
    "precio_liquidacion":   ["precio_liquidacion",  "precio liquidacion","liquidacion"],
}


def detectar_columna(headers: list[str], aliases: list[str]) -> Optional[int]:
    for alias in aliases:
        for i, h in enumerate(headers):
            if alias.lower() in str(h).lower():
                return i
    return None


@router.get("/actual")
async def lp_actual(
    marca_id:    Optional[int] = None,
    categoria_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
):
    params = {}
    where = []
    if marca_id:
        where.append("p.marca_id = :marca_id")
        params["marca_id"] = marca_id
    if categoria_id:
        where.append("p.categoria_id = :categoria_id")
        params["categoria_id"] = categoria_id
    where_sql = ("AND " + " AND ".join(where)) if where else ""
    rows = await db.execute(text(f"""
        SELECT p.sku, p.descripcion,
               m.nombre AS marca, c.nombre AS categoria,
               p.precio_venta_bruto, p.precio_venta_neto,
               p.costo_unitario_neto, p.precio_minimo_evento, p.precio_liquidacion,
               p.activo
        FROM productos p
        LEFT JOIN marcas m ON m.id = p.marca_id
        LEFT JOIN categorias c ON c.id = p.categoria_id
        WHERE p.activo = TRUE {where_sql}
        ORDER BY m.nombre, p.descripcion
    """), params)
    return [dict(r) for r in rows.mappings().all()]


@router.get("/historial")
async def lp_historial(db: AsyncSession = Depends(get_db)):
    rows = await db.execute(text("""
        SELECT id, nombre_archivo, descripcion, subido_por, subido_en,
               n_skus, n_actualizados, campos_actualizados, activa
        FROM lp_historial
        ORDER BY subido_en DESC
        LIMIT 50
    """))
    return [dict(r) for r in rows.mappings().all()]


@router.post("/upload", dependencies=[Depends(require_rol("admin", "editor"))])
async def upload_lp(
    file: UploadFile = File(...),
    descripcion: str = Form(""),
    usuario: str = Form(""),
    db: AsyncSession = Depends(get_db),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Solo se aceptan archivos Excel (.xlsx / .xls)")

    contenido = await file.read()
    wb = openpyxl.load_workbook(BytesIO(contenido), data_only=True)
    ws = wb.active

    # Detectar headers en fila 1
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    col_sku = None
    for i, h in enumerate(headers):
        if str(h).lower().strip() in ("sku", "cod", "codigo", "código"):
            col_sku = i
            break
    if col_sku is None:
        raise HTTPException(400, "No se encontró columna SKU en la primera fila del Excel")

    cols_precio = {}
    for campo, aliases in CAMPOS_PRECIO.items():
        idx = detectar_columna(headers, aliases)
        if idx is not None:
            cols_precio[campo] = idx

    if not cols_precio:
        raise HTTPException(400, f"No se detectó ninguna columna de precio. Headers encontrados: {headers[:10]}")

    # Cargar SKUs válidos
    skus_bd_rows = await db.execute(text("SELECT sku FROM productos WHERE activo=TRUE"))
    skus_bd = {r[0] for r in skus_bd_rows.fetchall()}

    actualizados = 0
    skus_procesados = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = str(row[col_sku]).strip() if row[col_sku] else ""
        if not sku or sku not in skus_bd:
            continue

        set_parts = []
        params = {"sku": sku}
        for campo, idx in cols_precio.items():
            val = row[idx] if idx < len(row) else None
            if val is not None and str(val).strip() not in ("", "-", "None"):
                try:
                    params[campo] = float(str(val).replace("$", "").replace(".", "").replace(",", "."))
                    set_parts.append(f"{campo} = :{campo}")
                except ValueError:
                    pass

        if set_parts:
            await db.execute(
                text(f"UPDATE productos SET {', '.join(set_parts)}, updated_at=NOW() WHERE sku=:sku"),
                params,
            )
            skus_procesados.add(sku)
            actualizados += 1

    await db.commit()

    # Marcar todas como no activa, luego esta como activa
    await db.execute(text("UPDATE lp_historial SET activa=FALSE"))
    await db.execute(text("""
        INSERT INTO lp_historial (nombre_archivo, descripcion, subido_por, n_skus, n_actualizados, campos_actualizados, activa)
        VALUES (:nombre, :desc, :user, :n_skus, :n_upd, :campos, TRUE)
    """), {
        "nombre": file.filename,
        "desc":   descripcion,
        "user":   usuario,
        "n_skus": len(skus_procesados),
        "n_upd":  actualizados,
        "campos": list(cols_precio.keys()),
    })
    await db.commit()

    return {
        "ok": True,
        "n_skus": len(skus_procesados),
        "n_actualizados": actualizados,
        "campos_detectados": list(cols_precio.keys()),
    }


@router.delete("/historial/{id}", dependencies=[Depends(require_rol("admin"))])
async def eliminar_historial(id: int, db: AsyncSession = Depends(get_db)):
    await db.execute(text("DELETE FROM lp_historial WHERE id=:id"), {"id": id})
    await db.commit()
    return {"ok": True}
