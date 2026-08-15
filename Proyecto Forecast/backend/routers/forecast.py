# routers/forecast.py

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_
from typing import List, Optional
from io import BytesIO
from database import get_db
from models.models import Forecast, Producto, Marca, Categoria, Subcategoria, Temporada
from schemas.schemas import ForecastCreate, ForecastUpdate, ForecastOut, ForecastPivotRow
from pydantic import BaseModel
from decimal import Decimal

import openpyxl

router = APIRouter()


@router.get("/", response_model=List[ForecastOut])
async def listar_forecast(
    sku: Optional[str] = None,
    anio: Optional[int] = None,
    temporada_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db)
):
    q = select(Forecast)
    if sku:          q = q.where(Forecast.sku == sku)
    if anio:         q = q.where(Forecast.anio == anio)
    if temporada_id: q = q.where(Forecast.temporada_id == temporada_id)
    result = await db.execute(q.order_by(Forecast.sku, Forecast.anio, Forecast.mes))
    return result.scalars().all()


@router.get("/pivot", response_model=List[ForecastPivotRow])
async def forecast_pivot(
    anio: int,
    temporada_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db)
):
    """Retorna forecast en formato pivot: un row por SKU con lista de 12 meses."""
    q = select(Forecast).where(Forecast.anio == anio)
    if temporada_id:
        q = q.where(Forecast.temporada_id == temporada_id)
    result = await db.execute(q.order_by(Forecast.sku, Forecast.mes))
    rows = result.scalars().all()

    pivot: dict[str, list[int]] = {}
    for r in rows:
        if r.sku not in pivot:
            pivot[r.sku] = [0] * 12
        pivot[r.sku][r.mes - 1] = r.cantidad

    return [ForecastPivotRow(sku=sku, anio=anio, meses=meses) for sku, meses in pivot.items()]


@router.post("/", response_model=ForecastOut, status_code=201)
async def crear_forecast(data: ForecastCreate, db: AsyncSession = Depends(get_db)):
    existente = await db.get(Forecast, {"sku": data.sku, "anio": data.anio, "mes": data.mes})
    if existente:
        raise HTTPException(400, "Ya existe un forecast para ese SKU/año/mes. Usa PUT para actualizar.")
    nuevo = Forecast(**data.model_dump())
    db.add(nuevo)
    await db.commit()
    await db.refresh(nuevo)
    return nuevo


@router.post("/bulk-upsert", response_model=List[ForecastOut])
async def bulk_upsert_forecast(items: List[ForecastCreate], db: AsyncSession = Depends(get_db)):
    """Carga masiva: inserta o actualiza múltiples registros de forecast."""
    resultado = []
    for item in items:
        q = select(Forecast).where(
            and_(Forecast.sku == item.sku, Forecast.anio == item.anio, Forecast.mes == item.mes)
        )
        existente = (await db.execute(q)).scalar_one_or_none()
        if existente:
            existente.cantidad = item.cantidad
            existente.temporada_id = item.temporada_id
            resultado.append(existente)
        else:
            nuevo = Forecast(**item.model_dump())
            db.add(nuevo)
            resultado.append(nuevo)
    await db.commit()
    for r in resultado:
        await db.refresh(r)
    return resultado


@router.put("/{id}", response_model=ForecastOut)
async def actualizar_forecast(id: int, data: ForecastUpdate, db: AsyncSession = Depends(get_db)):
    fc = await db.get(Forecast, id)
    if not fc:
        raise HTTPException(404, "Forecast no encontrado")
    fc.cantidad = data.cantidad
    await db.commit()
    await db.refresh(fc)
    return fc


@router.delete("/{id}", status_code=204)
async def eliminar_forecast(id: int, db: AsyncSession = Depends(get_db)):
    fc = await db.get(Forecast, id)
    if not fc:
        raise HTTPException(404, "Forecast no encontrado")
    await db.delete(fc)
    await db.commit()


# ─── Vista tabla pivot completa (para la UI spreadsheet) ─────────────────────

class TablaFilaOut(BaseModel):
    sku:                str
    marca:              Optional[str]
    tipo:               Optional[str]
    descripcion:        Optional[str]
    categoria:          Optional[str]
    subcategoria:       Optional[str]
    tipo_producto:      Optional[str]
    temporada:          Optional[str]
    precio_lp:          float
    precio_neto:        float
    forecast:           List[int]
    ventas_2025:        List[int]
    por_discontinuar:   bool = False
    activo:             bool = True
    comentario:         Optional[str] = None
    mes_agota_stock:    Optional[int] = None
    compras_necesarias: Optional[int] = None
    ajustado_stock_q4:  bool = False


@router.get("/tabla", response_model=List[TablaFilaOut])
async def forecast_tabla(
    anio: int = 2026,
    temporada_nombre: Optional[str] = None,
    db: AsyncSession = Depends(get_db)
):
    from sqlalchemy import text as sa_text

    # SKUs con forecast data en el año (para incluir inactivos que tengan datos)
    fc_skus_rows = (await db.execute(
        select(Forecast.sku).where(Forecast.anio == anio).distinct()
    )).scalars().all()
    skus_con_forecast = set(fc_skus_rows)

    # Productos activos + inactivos con forecast data
    q_prod = (
        select(Producto, Marca, Categoria, Subcategoria, Temporada)
        .join(Marca,         Producto.marca_id         == Marca.id,         isouter=True)
        .join(Categoria,     Producto.categoria_id     == Categoria.id,     isouter=True)
        .join(Subcategoria,  Producto.subcategoria_id  == Subcategoria.id,  isouter=True)
        .join(Temporada,     Producto.temporada_id     == Temporada.id,     isouter=True)
        .where(
            (Producto.activo == True) |
            (Producto.sku.in_(skus_con_forecast))
        )
    )
    if temporada_nombre:
        q_prod = q_prod.where(Temporada.nombre == temporada_nombre)
    filas_prod = (await db.execute(q_prod)).all()

    if not filas_prod:
        return []

    skus = [row[0].sku for row in filas_prod]

    # Forecast existente
    fc_rows = (await db.execute(
        select(Forecast).where(and_(Forecast.sku.in_(skus), Forecast.anio == anio))
    )).scalars().all()
    fc_map: dict[tuple, int] = {(f.sku, f.mes): f.cantidad for f in fc_rows}

    # Stock por SKU: base actual + llegadas con sus ETAs
    stock_rows = await db.execute(sa_text("""
        SELECT sku,
               COALESCE(stock_base,      0) +
               COALESCE(stock_full_ml,   0) +
               COALESCE(stock_full_fala, 0)  AS base,
               COALESCE(bodega_transito, 0)  AS bodega_transito,
               eta_transito,
               COALESCE(por_arribar,     0)  AS por_arribar,
               eta_arribar,
               COALESCE(pi,              0)  AS pi,
               eta_pi
        FROM stock
        WHERE sku = ANY(:skus)
    """), {"skus": skus})
    stock_map = {r["sku"]: dict(r) for r in stock_rows.mappings()}

    # Ventas reales 2025 por SKU y mes
    v2025_rows = await db.execute(sa_text("""
        SELECT sku,
               EXTRACT(MONTH FROM fecha)::int AS mes,
               SUM(cantidad) AS total
        FROM ventas
        WHERE EXTRACT(YEAR FROM fecha) = 2025
          AND sku = ANY(:skus)
        GROUP BY sku, mes
    """), {"skus": skus})
    v2025_map: dict[tuple, int] = {}
    for r in v2025_rows.mappings():
        v2025_map[(r["sku"], r["mes"])] = int(r["total"])

    def llegadas_por_mes(s: dict) -> dict:
        """Distribuye bodega_transito/por_arribar/pi al mes de su ETA para el año del forecast."""
        por_mes: dict[int, int] = {}
        for qty_key, eta_key in [
            ("bodega_transito", "eta_transito"),
            ("por_arribar",     "eta_arribar"),
            ("pi",              "eta_pi"),
        ]:
            qty = int(s.get(qty_key) or 0)
            eta = s.get(eta_key)
            if qty > 0 and eta:
                mes_eta = eta.month if hasattr(eta, "month") else None
                if mes_eta:
                    por_mes[mes_eta] = por_mes.get(mes_eta, 0) + qty
        return por_mes

    def calcular_mes_agota(sku: str, meses: list[int]) -> Optional[int]:
        """Simula consumo desde enero considerando stock actual + llegadas por ETA."""
        s = stock_map.get(sku)
        if not s:
            return 1
        stock = int(s["base"])
        llega = llegadas_por_mes(s)
        for mes in range(1, 13):
            stock += llega.get(mes, 0)
            stock -= meses[mes - 1]
            if stock <= 0:
                return mes
        return None

    def ajustar_forecast_q4(sku: str, meses: list[int], temporada_nombre: Optional[str]) -> bool:
        """
        La demanda proyectada Sep-Dic (índices 8-11) no puede superar el stock
        total disponible para venderlo en ese horizonte: bodega + Full ML +
        Full Falabella (ya en `base`) + en tránsito (ya zarpó) + por arribar
        (proforma). Si la supera, se escala proporcionalmente hacia abajo
        preservando la forma relativa entre los 4 meses. Afecta sobre todo a
        productos estacionales, que concentran demanda en este período.

        Excepción (indicación de negocio ago-2026): los productos "No
        Estacional" pueden reponerse en cualquier momento del año, así que
        no se topan al stock actual — se deja ver la demanda real proyectada
        y la faltante queda visible vía `compras_necesarias` (alerta de
        compra), en vez de esconderla recortando el forecast.
        """
        if (temporada_nombre or '').strip().lower() == 'no estacional':
            return False

        demanda_q4 = sum(meses[8:12])
        if demanda_q4 <= 0:
            return False
        s = stock_map.get(sku)
        if not s:
            disponible_q4 = 0
        else:
            disponible_q4 = (
                int(s["base"]) + int(s["bodega_transito"] or 0) + int(s["por_arribar"] or 0)
            )
        if demanda_q4 <= disponible_q4:
            return False

        factor = disponible_q4 / demanda_q4
        restante = disponible_q4
        for i in [8, 9, 10]:
            ajustado = int(meses[i] * factor)
            meses[i] = ajustado
            restante -= ajustado
        meses[11] = max(0, restante)  # el último mes absorbe el redondeo
        return True

    def calcular_compras_necesarias(sku: str, meses: list[int]) -> int:
        """Unidades a comprar para cubrir el forecast dado stock disponible + llegadas."""
        total_fc = sum(meses)
        if total_fc == 0:
            return 0
        s = stock_map.get(sku)
        base = int(s["base"]) if s else 0
        llegadas = sum(llegadas_por_mes(s).values()) if s else 0
        return max(0, total_fc - base - llegadas)

    resultado = []
    for producto, marca, categoria, subcategoria, temporada in filas_prod:
        meses = [fc_map.get((producto.sku, m), 0) for m in range(1, 13)]

        # Cero meses fuera de temporada usando fecha_inicio/fecha_fin de la temporada
        temp_nombre = (temporada.nombre or '').lower() if temporada else ''
        if temporada and temporada.fecha_inicio and temporada.fecha_fin:
            m_ini = temporada.fecha_inicio.month  # mes inicio (1-12)
            m_fin = temporada.fecha_fin.month     # mes fin (1-12)
            if m_ini <= m_fin:
                meses_temporada = set(range(m_ini, m_fin + 1))
            else:  # cruza año (ej: Nov-Feb)
                meses_temporada = set(range(m_ini, 13)) | set(range(1, m_fin + 1))
            for m in range(1, 13):
                if m not in meses_temporada:
                    meses[m - 1] = 0
        elif 'invierno' in temp_nombre:
            # Fallback si no hay fechas: Invierno = Mar-Ago
            for i in [8, 9, 10, 11, 0, 1]:
                meses[i] = 0
        elif 'verano' in temp_nombre:
            # Fallback si no hay fechas: Verano = Sep-Feb
            for i in [2, 3, 4, 5, 6, 7]:
                meses[i] = 0

        por_disc = bool(getattr(producto, 'por_discontinuar', False))
        mes_agota = None
        compras_nec = None

        if por_disc:
            mes_agota = calcular_mes_agota(producto.sku, meses)
            if mes_agota is not None:
                for i in range(mes_agota, 12):
                    meses[i] = 0
        else:
            # Solo calcular compras necesarias para productos sin temporada definida
            es_estacional = temporada and (
                (temporada.fecha_inicio and temporada.fecha_fin) or
                'invierno' in (temporada.nombre or '').lower() or
                'verano'   in (temporada.nombre or '').lower()
            )
            if not es_estacional:
                cn = calcular_compras_necesarias(producto.sku, meses)
                compras_nec = cn if cn > 0 else None

        ajustado_q4 = ajustar_forecast_q4(producto.sku, meses, temporada.nombre if temporada else None)

        precio_neto = float(producto.precio_venta_neto or 0) or round(float(producto.precio_venta_bruto or 0) / 1.19, 2)

        v2025 = [v2025_map.get((producto.sku, m), 0) for m in range(1, 13)]

        resultado.append(TablaFilaOut(
            sku                = producto.sku,
            marca              = marca.nombre        if marca        else None,
            tipo               = temporada.nombre    if temporada    else None,
            descripcion        = producto.descripcion,
            categoria          = categoria.nombre    if categoria    else None,
            subcategoria       = subcategoria.nombre if subcategoria else None,
            tipo_producto      = producto.tipo_producto,
            temporada          = temporada.nombre    if temporada    else None,
            precio_lp          = float(producto.precio_venta_bruto or 0),
            precio_neto        = precio_neto,
            forecast           = meses,
            ventas_2025        = v2025,
            por_discontinuar   = por_disc,
            activo             = producto.activo,
            comentario         = producto.comentario,
            mes_agota_stock    = mes_agota,
            compras_necesarias = compras_nec,
            ajustado_stock_q4  = ajustado_q4,
        ))

    resultado.sort(key=lambda r: (r.marca or '', r.descripcion or ''))
    return resultado


# ─── Proyeccion Q4 2026 (Oct-Nov-Dic) ────────────────────────────────────────

@router.get("/proyeccion-q4")
async def proyeccion_q4(db: AsyncSession = Depends(get_db)):
    """
    Retorna las proyecciones modelo ANCLA-SI-MACRO para Oct/Nov/Dic 2026
    como dict {sku: {10: qty, 11: qty, 12: qty}}.
    """
    from sqlalchemy import text
    rows = await db.execute(text(
        "SELECT sku, mes, cantidad FROM proyeccion_q4_2026 ORDER BY sku, mes"
    ))
    result: dict = {}
    for r in rows.mappings():
        result.setdefault(r['sku'], {})[r['mes']] = r['cantidad']
    return result


class ProyeccionQ4Item(BaseModel):
    sku:      str
    mes:      int   # 10, 11 o 12
    cantidad: int

@router.post("/proyeccion-q4/upsert")
async def upsert_proyeccion_q4(items: List[ProyeccionQ4Item], db: AsyncSession = Depends(get_db)):
    """Guarda ediciones manuales en proyeccion_q4_2026."""
    from sqlalchemy import text
    for item in items:
        if item.mes not in (10, 11, 12):
            continue
        await db.execute(text("""
            INSERT INTO proyeccion_q4_2026 (sku, mes, cantidad)
            VALUES (:sku, :mes, :cantidad)
            ON CONFLICT (sku, mes) DO UPDATE SET cantidad = EXCLUDED.cantidad, updated_at = NOW()
        """), {"sku": item.sku, "mes": item.mes, "cantidad": item.cantidad})
    await db.commit()
    return {"ok": True, "updated": len(items)}


import asyncio, sys, os

@router.post("/proyeccion-q4/recalcular", dependencies=[Depends(lambda: None)])
async def recalcular_proyeccion_q4():
    """Recalcula las proyecciones Q4 2026 ejecutando el script ANCLA-SI-MACRO."""
    script = os.path.normpath(
        os.path.join(os.path.dirname(__file__), '..', 'proyectar_q4_2026.py')
    )
    if not os.path.exists(script):
        raise HTTPException(500, "proyectar_q4_2026.py no encontrado")
    try:
        proc = await asyncio.create_subprocess_exec(
            sys.executable, script,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.STDOUT,
            cwd=os.path.dirname(script),
        )
        try:
            stdout, _ = await asyncio.wait_for(proc.communicate(), timeout=60)
        except asyncio.TimeoutError:
            proc.kill()
            await proc.communicate()
            raise HTTPException(504, "Timeout al proyectar Q4")
        salida = (stdout or b"").decode("utf-8", errors="replace")
        return {"ok": proc.returncode == 0, "salida": salida[-2000:]}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


# ─── Toggle por_discontinuar ──────────────────────────────────────────────────

class DiscontinuarPayload(BaseModel):
    sku:             str
    por_discontinuar: bool

@router.post("/discontinuar")
async def set_discontinuar(payload: DiscontinuarPayload, db: AsyncSession = Depends(get_db)):
    from sqlalchemy import text
    await db.execute(text(
        "UPDATE productos SET por_discontinuar = :val WHERE sku = :sku"
    ), {"val": payload.por_discontinuar, "sku": payload.sku})
    await db.commit()
    return {"ok": True, "sku": payload.sku, "por_discontinuar": payload.por_discontinuar}


# ─── Forecast automático Holt-Winters ────────────────────────────────────────

class HoltWintersResult(BaseModel):
    sku: str
    meses_generados: int
    metodo: str
    datos_historicos_meses: int
    forecast: list  # [{anio, mes, cantidad}]

@router.post("/generar-holt-winters/{sku}", response_model=HoltWintersResult)
async def generar_holt_winters(
    sku: str,
    desde_mes: int = 7,
    desde_anio: int = 2026,
    hasta_mes: int = 12,
    hasta_anio: int = 2026,
    db: AsyncSession = Depends(get_db),
):
    from sqlalchemy import text
    import pandas as pd
    import numpy as np

    # 1. Verificar que el producto existe
    prod = await db.get(Producto, sku)
    if not prod:
        raise HTTPException(404, f"Producto {sku} no encontrado")

    # 2. Traer ventas netas mensuales históricas (cantidad - devoluciones)
    rows = await db.execute(text("""
        SELECT
            EXTRACT(YEAR  FROM fecha)::int AS anio,
            EXTRACT(MONTH FROM fecha)::int AS mes,
            SUM(cantidad - COALESCE(unidades_devueltas, 0)) AS uds
        FROM ventas
        WHERE sku = :sku
          AND (estado_orden IS NULL OR estado_orden != 'Cancelado')
          AND activo = true
          AND fecha <= CURRENT_DATE
        GROUP BY 1, 2
        ORDER BY 1, 2
    """), {"sku": sku})
    ventas_raw = rows.mappings().all()

    if not ventas_raw:
        raise HTTPException(422, f"SKU {sku} sin historial de ventas — no se puede proyectar")

    # 3. Construir serie mensual completa (rellenar meses sin ventas con 0)
    df = pd.DataFrame([dict(r) for r in ventas_raw])
    df["anio"] = df["anio"].astype(int)
    df["mes"]  = df["mes"].astype(int)
    df["uds"]  = df["uds"].astype(float).clip(lower=0)

    anio_min, mes_min = int(df["anio"].min()), int(df["mes"].min())
    anio_max, mes_max = int(df["anio"].max()), int(df["mes"].max())

    # Índice mensual completo
    idx_inicio = pd.Period(f"{anio_min}-{mes_min:02d}", freq="M")
    idx_fin    = pd.Period(f"{anio_max}-{mes_max:02d}", freq="M")
    idx_completo = pd.period_range(idx_inicio, idx_fin, freq="M")

    serie = pd.Series(0.0, index=idx_completo)
    for _, r in df.iterrows():
        p = pd.Period(f"{int(r.anio)}-{int(r.mes):02d}", freq="M")
        if p in serie.index:
            serie[p] = float(r.uds)

    n_meses = len(serie)

    # 4. Construir lista de meses a proyectar
    meses_objetivo = []
    a, m = desde_anio, desde_mes
    while (a, m) <= (hasta_anio, hasta_mes):
        meses_objetivo.append((a, m))
        m += 1
        if m > 12:
            m = 1
            a += 1
    n_proyectar = len(meses_objetivo)

    # 5. Elegir método según cantidad de datos históricos
    if n_meses >= 24:
        metodo = "holt_winters"
        from statsmodels.tsa.holtwinters import ExponentialSmoothing
        model = ExponentialSmoothing(
            serie.values,
            trend="add",
            seasonal="add",
            seasonal_periods=12,
            initialization_method="estimated",
        )
        fit = model.fit(optimized=True)
        # Calcular pasos usando ordinals (Period subtraction da MonthEnd, no int)
        ultimo     = pd.Period(f"{anio_max}-{mes_max:02d}", freq="M")
        primer_obj = pd.Period(f"{desde_anio}-{desde_mes:02d}", freq="M")
        hasta_obj  = pd.Period(f"{hasta_anio}-{hasta_mes:02d}", freq="M")
        pasos_hasta = hasta_obj.ordinal - ultimo.ordinal   # cuántos pasos forecaster desde el fin del historial
        pasos_desde = primer_obj.ordinal - ultimo.ordinal  # pasos hasta el inicio solicitado

        if pasos_hasta <= 0:
            # El historial ya cubre el período pedido: usar los valores reales como proyección
            valores = []
            for (a, m) in meses_objetivo:
                p = pd.Period(f"{a}-{m:02d}", freq="M")
                if p in serie.index:
                    valores.append(float(serie[p]))
                else:
                    valores.append(float(serie.mean()))
        else:
            raw = fit.forecast(pasos_hasta)  # array de pasos_hasta valores
            if pasos_desde <= 0:
                # El inicio solicitado está dentro o antes del historial
                valores = list(raw[:n_proyectar])
            else:
                # Índice 0-based: paso 1 = primer mes tras ultimo
                valores = list(raw[pasos_desde - 1: pasos_desde - 1 + n_proyectar])

    elif n_meses >= 12:
        metodo = "promedio_estacional"
        # Promedio por mes del año usando todos los datos disponibles
        por_mes = {i: [] for i in range(1, 13)}
        for period, val in serie.items():
            por_mes[period.month].append(val)
        promedios = {m_: float(np.mean(v)) if v else 0.0 for m_, v in por_mes.items()}
        valores = [promedios[m] for (_, m) in meses_objetivo]

    else:
        metodo = "promedio_simple"
        media = float(serie.mean())
        valores = [media] * n_proyectar

    # 6. Upsert: meses Q4 (oct-dic) → proyeccion_q4_2026 (O PROY), resto → forecast.cantidad
    forecast_result = []
    for (a, m), val in zip(meses_objetivo, valores):
        cantidad = max(0, int(round(val)))
        if m in (10, 11, 12) and a == 2026:
            await db.execute(text("""
                INSERT INTO proyeccion_q4_2026 (sku, mes, cantidad)
                VALUES (:sku, :mes, :cantidad)
                ON CONFLICT (sku, mes) DO UPDATE SET cantidad = EXCLUDED.cantidad, updated_at = NOW()
            """), {"sku": sku, "mes": m, "cantidad": cantidad})
        else:
            await db.execute(text("""
                INSERT INTO forecast (sku, anio, mes, cantidad)
                VALUES (:sku, :anio, :mes, :cantidad)
                ON CONFLICT (sku, anio, mes)
                DO UPDATE SET cantidad = EXCLUDED.cantidad, updated_at = NOW()
            """), {"sku": sku, "anio": a, "mes": m, "cantidad": cantidad})
        forecast_result.append({"anio": a, "mes": m, "cantidad": cantidad})

    await db.commit()

    return HoltWintersResult(
        sku=sku,
        meses_generados=n_proyectar,
        metodo=metodo,
        datos_historicos_meses=n_meses,
        forecast=forecast_result,
    )


# ─── Análisis Stock-Constrained por SKU ──────────────────────────────────────

class StockAnalisisOut(BaseModel):
    sku:                str
    descripcion:        Optional[str]
    precio_lp:          float
    stock_actual:       int          # stock físico ahora
    mes_actual:         int          # mes corriente (1-12)
    forecast:           List[int]    # 12 meses, índice 0=Ene
    ventas_reales:      List[int]    # ventas 2026 reales por mes
    llegadas:           List[int]    # llegadas programadas por mes (índice 0=Ene)
    lo_proyectado:      List[int]    # forecast stock-constrained
    compra_para_fc:     int          # unidades a comprar para cubrir forecast
    compra_para_proy:   int          # unidades a comprar para cubrir proyección

@router.get("/stock-analisis/{sku}", response_model=StockAnalisisOut)
async def stock_analisis(sku: str, anio: int = 2026, db: AsyncSession = Depends(get_db)):
    from sqlalchemy import text
    import datetime

    mes_actual = datetime.date.today().month  # mes corriente

    # Producto
    prod = (await db.execute(
        select(Producto).where(Producto.sku == sku)
    )).scalar_one_or_none()
    if not prod:
        raise HTTPException(404, f"SKU {sku} no encontrado")

    # Forecast
    fc_rows = (await db.execute(
        select(Forecast).where(Forecast.sku == sku, Forecast.anio == anio)
    )).scalars().all()
    forecast = [0] * 12
    for f in fc_rows:
        forecast[f.mes - 1] = f.cantidad

    # Ventas reales 2026 por mes
    v_rows = await db.execute(text("""
        SELECT EXTRACT(MONTH FROM fecha)::int AS mes, SUM(cantidad)::int AS total
        FROM ventas
        WHERE sku = :sku AND EXTRACT(YEAR FROM fecha) = :anio
        GROUP BY mes
    """), {"sku": sku, "anio": anio})
    ventas_reales = [0] * 12
    for r in v_rows.mappings():
        ventas_reales[r["mes"] - 1] = r["total"]

    # Stock y llegadas por mes
    s_row = (await db.execute(text("""
        SELECT
            COALESCE(stock_base, 0) + COALESCE(stock_full_ml, 0) + COALESCE(stock_full_fala, 0) AS stock_actual,
            COALESCE(llegada_jun, 0) AS lleg_6,
            COALESCE(llegada_jul, 0) AS lleg_7,
            COALESCE(llegada_ago, 0) AS lleg_8,
            COALESCE(llegada_sep, 0) AS lleg_9,
            COALESCE(llegada_oct, 0) AS lleg_10,
            COALESCE(llegada_nov, 0) AS lleg_11,
            COALESCE(llegada_dic, 0) AS lleg_12
        FROM stock WHERE sku = :sku
    """), {"sku": sku})).mappings().first()

    stock_actual = int(s_row["stock_actual"]) if s_row else 0
    llegadas = [0] * 12  # índice 0=Ene
    if s_row:
        for col, mes in [("lleg_6",6),("lleg_7",7),("lleg_8",8),("lleg_9",9),
                         ("lleg_10",10),("lleg_11",11),("lleg_12",12)]:
            llegadas[mes - 1] = int(s_row[col])

    # Calcular Lo Proyectado mes a mes
    lo_proyectado = [0] * 12
    stock_disponible = stock_actual

    for m in range(1, 13):  # m = mes (1-12)
        i = m - 1
        lleg = llegadas[i]
        stock_disponible += lleg

        if m < mes_actual:
            # Pasado: usar ventas reales
            lo_proyectado[i] = ventas_reales[i]
            # No descontar stock (ya ocurrió)
        elif m == mes_actual:
            # Mes actual: lo ya vendido + lo que queda en stock
            lo_proyectado[i] = ventas_reales[i] + max(0, stock_disponible)
            stock_disponible = 0  # todo el stock se usa en el mes actual
        else:
            # Futuro: limitado por stock disponible
            puede_vender = min(forecast[i], max(0, stock_disponible))
            lo_proyectado[i] = puede_vender
            stock_disponible -= puede_vender

    # Compras necesarias a partir del mes siguiente al actual
    meses_futuros = list(range(mes_actual, 12))  # índices 0-based para mes_actual+1 a Dic

    total_fc_futuro = sum(forecast[i] for i in meses_futuros)
    total_proy_futuro = sum(lo_proyectado[i] for i in meses_futuros)
    total_llegadas_futuras = sum(llegadas[i] for i in meses_futuros)
    # Stock disponible al inicio del próximo mes = 0 (se agota en mes actual)
    stock_para_futuro = max(0, stock_actual - sum(ventas_reales[mes_actual - 1:mes_actual]))
    # Simplificado: compras = necesidad futura - stock disponible - llegadas futuras
    compra_para_fc   = max(0, total_fc_futuro   - stock_para_futuro - total_llegadas_futuras)
    compra_para_proy = max(0, total_proy_futuro - stock_para_futuro - total_llegadas_futuras)

    return StockAnalisisOut(
        sku             = sku,
        descripcion     = prod.descripcion,
        precio_lp       = float(prod.precio_venta_bruto or 0),
        stock_actual    = stock_actual,
        mes_actual      = mes_actual,
        forecast        = forecast,
        ventas_reales   = ventas_reales,
        llegadas        = llegadas,
        lo_proyectado   = lo_proyectado,
        compra_para_fc  = compra_para_fc,
        compra_para_proy= compra_para_proy,
    )


# ─── Carga masiva desde Excel ─────────────────────────────────────────────────
MESES_HEADER = ['ene','feb','mar','abr','may','jun','jul','ago','sep','oct','nov','dic']

class CargaExcelResultado(BaseModel):
    procesados:  int
    creados:     int
    actualizados: int
    sin_cambio:  int
    errores:     List[str]


@router.post("/carga-excel", response_model=CargaExcelResultado)
async def carga_forecast_excel(
    anio: int = 2026,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db)
):
    """
    Carga masiva de forecast desde la plantilla Excel oficial.
    Formato: hoja 'Forecast {anio}', fila 1=título, fila 2=grupos, fila 3=cabeceras,
    filas 4+ = datos (col A=SKU, col B=Descripción, col C=Temporada, cols D-O=meses Ene-Dic).
    Ignora filas con SKU vacío o que empiezen con 'TOTAL'.
    """
    if not file.filename.endswith(('.xlsx', '.xlsm')):
        raise HTTPException(400, "El archivo debe ser .xlsx")

    contenido = await file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(contenido), data_only=True)
    except Exception:
        raise HTTPException(400, "No se pudo leer el archivo Excel. Verifica que no esté dañado.")

    # Buscar la hoja correcta
    hoja_nombre = f'Forecast {anio}'
    if hoja_nombre not in wb.sheetnames:
        # Intentar con la primera hoja
        hoja_nombre = wb.sheetnames[0]

    ws = wb[hoja_nombre]

    # Detectar fila de cabeceras buscando la celda "SKU" en las primeras 5 filas
    fila_header = None
    for row in ws.iter_rows(min_row=1, max_row=5):
        for c in row:
            if str(c.value or '').strip().upper() == 'SKU':
                fila_header = c.row
                break
        if fila_header:
            break

    if not fila_header:
        raise HTTPException(400, "No se encontró la fila de cabeceras (columna 'SKU') en las primeras 5 filas.")

    # Obtener SKUs válidos en la BD
    skus_bd = set(
        r[0] for r in (await db.execute(select(Producto.sku))).all()
    )

    errores  = []
    items    = []
    fila_data = fila_header + 1

    for row in ws.iter_rows(min_row=fila_data, values_only=True):
        sku = str(row[0] or '').strip()
        if not sku or sku.upper().startswith('TOTAL'):
            continue

        if sku not in skus_bd:
            errores.append(f"SKU '{sku}' no existe en el catálogo DCIC — fila ignorada")
            continue

        for mes_idx, col_offset in enumerate(range(3, 15)):   # cols D(3) a O(14), 0-based
            try:
                valor = row[col_offset]
                if valor is None:
                    cantidad = 0
                else:
                    cantidad = int(float(str(valor).replace(',', '').strip()))
                if cantidad < 0:
                    errores.append(f"SKU '{sku}' mes {mes_idx+1}: valor negativo ({cantidad}) → se usa 0")
                    cantidad = 0
                items.append({'sku': sku, 'anio': anio, 'mes': mes_idx + 1, 'cantidad': cantidad})
            except (ValueError, TypeError):
                errores.append(f"SKU '{sku}' mes {mes_idx+1}: valor no numérico → se usa 0")
                items.append({'sku': sku, 'anio': anio, 'mes': mes_idx + 1, 'cantidad': 0})

    if not items:
        raise HTTPException(400, "No se encontraron datos válidos en el archivo.")

    # Upsert
    creados = actualizados = sin_cambio = 0

    for item in items:
        q = select(Forecast).where(
            and_(Forecast.sku  == item['sku'],
                 Forecast.anio == item['anio'],
                 Forecast.mes  == item['mes'])
        )
        existente = (await db.execute(q)).scalar_one_or_none()

        if existente:
            if existente.cantidad != item['cantidad']:
                existente.cantidad = item['cantidad']
                actualizados += 1
            else:
                sin_cambio += 1
        else:
            db.add(Forecast(sku=item['sku'], anio=item['anio'],
                            mes=item['mes'], cantidad=item['cantidad']))
            creados += 1

    await db.commit()

    return CargaExcelResultado(
        procesados=len(items),
        creados=creados,
        actualizados=actualizados,
        sin_cambio=sin_cambio,
        errores=errores,
    )
