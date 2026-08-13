"""
Script puntual: carga stock_full_ml (reporte Full Mercado Libre) y
stock_full_fala (reporte Full FBF Falabella) desde los Excel entregados.

  - ML: hoja "Resumen", SKU en col D, stock = "Aptas para vender" (col S)
  - Falabella FBF: hoja "Product details", SKU = "SKU Vendedor", stock = "Stock Disponible"
"""
import openpyxl
from fastapi.testclient import TestClient
from main import app

ML_PATH   = r"C:\Users\rafae\OneDrive\Escritorio\Proyecto Forecast\stock_general_full_12124832_641896fd6fa0c9a3f66595c4a60f59ac.xlsx"
FALA_PATH = r"C:\Users\rafae\OneDrive\Escritorio\Proyecto Forecast\Product_details_Full FBF.xlsx"

# ── Leer ML ───────────────────────────────────────────────────────────────────
wb_ml = openpyxl.load_workbook(ML_PATH, data_only=True)
ws_ml = wb_ml["Resumen"]
stock_ml = {}
for row in ws_ml.iter_rows(min_row=14, values_only=True):
    sku = row[3]
    aptas = row[18]
    if sku:
        stock_ml[sku] = int(aptas or 0)
print(f"ML: {len(stock_ml)} SKUs leídos")

# ── Leer Falabella FBF ────────────────────────────────────────────────────────
wb_fa = openpyxl.load_workbook(FALA_PATH, data_only=True)
ws_fa = wb_fa["Product details"]
stock_fala = {}
for row in ws_fa.iter_rows(min_row=2, values_only=True):
    sku = row[4]
    disponible = row[12]
    if sku:
        stock_fala[sku] = int(disponible or 0)
print(f"Falabella FBF: {len(stock_fala)} SKUs leídos")

# ── Combinar por SKU ───────────────────────────────────────────────────────────
todos_skus = set(stock_ml) | set(stock_fala)
items = []
for sku in todos_skus:
    item = {"sku": sku}
    if sku in stock_ml:
        item["stock_full_ml"] = stock_ml[sku]
    if sku in stock_fala:
        item["stock_full_fala"] = stock_fala[sku]
    items.append(item)

print(f"Total combinado: {len(items)} SKUs únicos")

# ── Enviar en lotes ────────────────────────────────────────────────────────────
BATCH = 200
resultados = {"actualizados": 0, "ignorados": 0, "desactivados": [], "reactivados": []}

with TestClient(app) as client:
    for i in range(0, len(items), BATCH):
        lote = items[i:i + BATCH]
        r = client.post("/api/stock/sync-full", json=lote)
        r.raise_for_status()
        data = r.json()
        resultados["actualizados"] += data.get("actualizados", 0)
        resultados["ignorados"]    += data.get("ignorados", 0)
        resultados["desactivados"].extend(data.get("desactivados", []))
        resultados["reactivados"].extend(data.get("reactivados", []))

print("=" * 60)
print(f"Actualizados: {resultados['actualizados']}")
print(f"Ignorados (SKU no existe en productos): {resultados['ignorados']}")
print(f"Auto-desactivados por stock=0: {len(resultados['desactivados'])} -> {resultados['desactivados'][:20]}")
print(f"Auto-marcados por_discontinuar: {len(resultados['reactivados'])} -> {resultados['reactivados'][:20]}")
