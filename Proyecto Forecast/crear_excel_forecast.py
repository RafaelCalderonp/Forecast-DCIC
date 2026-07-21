from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ── Colores / estilos ──────────────────────────────────────────────────────────
AZUL_OSC   = '1F3864'
AZUL_MED   = '2E75B6'
AZUL_CLAR  = 'D6E4F0'
VERDE_OSC  = '375623'
VERDE_CLAR = 'E2EFDA'
AMBAR      = 'C55A11'
AMBAR_CLAR = 'FCE4D6'
GRIS       = 'F2F2F2'
BLANCO     = 'FFFFFF'
AMARILLO   = 'FFF2CC'
ROJO_CLAR  = 'FCE4E4'

thin  = Side(style='thin',   color='BFBFBF')
thick = Side(style='medium', color='1F3864')
borde_thin  = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
borde_mes   = Border(left=thick, right=thin,  top=thin,  bottom=thin)

def cell(ws, row, col, val='', bold=False, bg=BLANCO, fg='000000',
         size=10, wrap=False, center=False, right=False, border=borde_thin, italic=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name='Arial', bold=bold, color=fg, size=size, italic=italic)
    c.fill      = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(
        horizontal='center' if center else ('right' if right else 'left'),
        vertical='center', wrap_text=wrap)
    c.border    = border
    return c

MESES = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
MESES_FULL = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
              'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']

TEMPORADAS_VALIDAS = ['Verano','Invierno','No Estacional','Verano/Rotativo']

# ══════════════════════════════════════════════════════════════════════════════
# HOJA 1 – Plantilla de carga  (la que el usuario llena)
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = 'Forecast 2026'
ws.freeze_panes = 'D4'
ws.sheet_view.showGridLines = False

# ── Título ─────────────────────────────────────────────────────────────────
ws.merge_cells('A1:P1')
c = ws['A1']
c.value = 'FORECAST 2026  –  DCIC  |  Una fila por SKU  |  Cantidades en unidades'
c.font  = Font(name='Arial', bold=True, size=13, color=BLANCO)
c.fill  = PatternFill('solid', fgColor=AZUL_OSC)
c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 28

# ── Fila de grupos de meses (trimestres) ────────────────────────────────────
grupos = [('Q1', 4, 6, AZUL_MED), ('Q2', 7, 9, '17375E'),
          ('Q3', 10, 12, AZUL_MED), ('Q4', 13, 15, '17375E')]
for label, c1, c2, bg in grupos:
    ws.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
    c_obj = ws.cell(row=2, column=c1, value=label)
    c_obj.font      = Font(name='Arial', bold=True, size=9, color=BLANCO)
    c_obj.fill      = PatternFill('solid', fgColor=bg)
    c_obj.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 14

# ── Cabeceras columnas (fila 3 — sin merge en fila 2 para A/B/C) ─────────────
HEADERS = [
    (1, 8,  'SKU',        True),
    (2, 28, 'Descripcion',False),
    (3, 18, 'Temporada',  True),
]
# Rellenar fila 2 para cols A-C con mismo color que fila 3
for col_n, width, label, is_key in HEADERS:
    ws.column_dimensions[get_column_letter(col_n)].width = width
    bg_hdr = AZUL_OSC if is_key else AZUL_MED
    # fila 2 (vacía pero misma paleta)
    c2 = ws.cell(row=2, column=col_n)
    c2.fill = PatternFill('solid', fgColor=bg_hdr)
    # fila 3 (etiqueta real)
    c = ws.cell(row=3, column=col_n, value=label)
    c.font      = Font(name='Arial', bold=True, color=BLANCO, size=10)
    c.fill      = PatternFill('solid', fgColor=bg_hdr)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border    = borde_thin

# Meses
TEMPORADA_COLORES = {
    1:AZUL_CLAR, 2:AZUL_CLAR,                          # Ene-Feb  (Verano)
    3:VERDE_CLAR, 4:VERDE_CLAR, 5:VERDE_CLAR,           # Mar-May  (Invierno)
    6:VERDE_CLAR, 7:VERDE_CLAR, 8:VERDE_CLAR,           # Jun-Ago  (Invierno)
    9:AZUL_CLAR, 10:AZUL_CLAR, 11:AZUL_CLAR, 12:AZUL_CLAR,  # Sep-Dic (Verano siguiente)
}
for i, mes in enumerate(MESES):
    col = i + 4
    ws.column_dimensions[get_column_letter(col)].width = 7
    c = ws.cell(row=3, column=col, value=mes)
    c.font      = Font(name='Arial', bold=True, color=BLANCO, size=9)
    c.fill      = PatternFill('solid', fgColor=AZUL_MED if i < 8 else '17375E')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border    = borde_thin

# Columna TOTAL
col_total = 16
ws.column_dimensions[get_column_letter(col_total)].width = 10
# fila 2
c2t = ws.cell(row=2, column=col_total)
c2t.fill = PatternFill('solid', fgColor=AZUL_OSC)
# fila 3
c = ws.cell(row=3, column=col_total, value='TOTAL')
c.font      = Font(name='Arial', bold=True, color=BLANCO, size=10)
c.fill      = PatternFill('solid', fgColor=AZUL_OSC)
c.alignment = Alignment(horizontal='center', vertical='center')
c.border    = borde_thin
ws.row_dimensions[3].height = 22

# ── Validación dropdown Temporada ────────────────────────────────────────────
dv = DataValidation(
    type='list',
    formula1='"Verano,Invierno,No Estacional,Verano/Rotativo"',
    allow_blank=False,
    showErrorMessage=True,
    errorTitle='Temporada inválida',
    error='Use: Verano / Invierno / No Estacional / Verano/Rotativo'
)
ws.add_data_validation(dv)
dv.sqref = 'C4:C500'

# ── Filas de ejemplo (5 SKUs) ────────────────────────────────────────────────
ejemplos = [
    ('DCX-001', 'Tabla SUP 10.6 Aqua',       'Verano',
     [0,0,0,0,0,0,0,0,150,120,200,180]),
    ('DCX-045', 'Chaleco Neopreno M Adulto',  'Invierno',
     [0,0,80,90,100,110,95,85,0,0,0,0]),
    ('DCX-112', 'Kayak Inflable 2P',          'No Estacional',
     [30,25,35,40,45,40,38,42,50,55,60,50]),
    ('DCX-200', 'Aleta Bodyboard Pro',        'Verano/Rotativo',
     [60,50,30,20,15,12,18,25,45,55,70,65]),
    ('DCX-310', 'Gafas Buceo Silicona',       'No Estacional',
     [20,18,22,24,25,23,22,24,28,30,32,26]),
]

for r_idx, (sku, desc, temp, meses_vals) in enumerate(ejemplos):
    row = r_idx + 4
    bg  = GRIS if r_idx % 2 == 0 else BLANCO

    cell(ws, row, 1, sku,  bold=True, bg=bg, fg='1F3864')
    cell(ws, row, 2, desc, bg=bg)
    cell(ws, row, 3, temp, bg=AMBAR_CLAR, fg=AMBAR, bold=True, center=True)

    total = sum(meses_vals)
    for i, v in enumerate(meses_vals):
        col = i + 4
        c = ws.cell(row=row, column=col, value=v)
        c.font      = Font(name='Arial', size=10, color='000000' if v > 0 else 'BBBBBB')
        c.fill      = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = borde_thin
        c.number_format = '#,##0'

    # Fórmula TOTAL
    c_tot = ws.cell(row=row, column=col_total)
    c_tot.value        = f'=SUM(D{row}:O{row})'
    c_tot.font         = Font(name='Arial', bold=True, size=10, color='1F3864')
    c_tot.fill         = PatternFill('solid', fgColor=AZUL_CLAR)
    c_tot.alignment    = Alignment(horizontal='center', vertical='center')
    c_tot.border       = borde_thin
    c_tot.number_format = '#,##0'

    ws.row_dimensions[row].height = 18

# ── Fila TOTAL general (fila 9) ──────────────────────────────────────────────
row_tot = len(ejemplos) + 4
ws.merge_cells(f'A{row_tot}:C{row_tot}')
c = ws.cell(row=row_tot, column=1, value='TOTAL FORECAST')
c.font      = Font(name='Arial', bold=True, size=10, color=BLANCO)
c.fill      = PatternFill('solid', fgColor=AZUL_OSC)
c.alignment = Alignment(horizontal='center', vertical='center')
c.border    = borde_thin

for i in range(12):
    col = i + 4
    col_ltr = get_column_letter(col)
    c = ws.cell(row=row_tot, column=col)
    c.value        = f'=SUM({col_ltr}4:{col_ltr}{row_tot-1})'
    c.font         = Font(name='Arial', bold=True, size=10, color=BLANCO)
    c.fill         = PatternFill('solid', fgColor=AZUL_OSC)
    c.alignment    = Alignment(horizontal='center', vertical='center')
    c.border       = borde_thin
    c.number_format = '#,##0'

c_tot2 = ws.cell(row=row_tot, column=col_total)
c_tot2.value        = f'=SUM(P4:P{row_tot-1})'
c_tot2.font         = Font(name='Arial', bold=True, size=10, color=BLANCO)
c_tot2.fill         = PatternFill('solid', fgColor=AZUL_OSC)
c_tot2.alignment    = Alignment(horizontal='center', vertical='center')
c_tot2.border       = borde_thin
c_tot2.number_format = '#,##0'
ws.row_dimensions[row_tot].height = 20

# ── Nota al pie ──────────────────────────────────────────────────────────────
row_nota = row_tot + 2
ws.merge_cells(f'A{row_nota}:P{row_nota}')
c = ws.cell(row=row_nota, column=1,
    value='  Instrucciones: (1) No modificar las columnas A-C ni los encabezados.  '
          '(2) Ingresar cantidades enteras >= 0.  '
          '(3) Dejar en 0 los meses sin venta (no borrar la celda).  '
          '(4) SKU debe coincidir exactamente con el catálogo DCIC.')
c.font      = Font(name='Arial', italic=True, size=9, color='595959')
c.fill      = PatternFill('solid', fgColor=AMARILLO)
c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws.row_dimensions[row_nota].height = 28


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 2 – Referencia de temporadas
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Referencia Temporadas')
ws2.sheet_view.showGridLines = False

ws2.merge_cells('A1:F1')
c = ws2['A1']
c.value = 'REFERENCIA: TIPOS DE TEMPORADA Y MESES ACTIVOS'
c.font  = Font(name='Arial', bold=True, size=13, color=BLANCO)
c.fill  = PatternFill('solid', fgColor=AZUL_OSC)
c.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 26

headers2 = ['Temporada', 'Meses activos', 'Meses con 0', 'Lead time', 'Descripción', 'Color en plantilla']
widths2  = [20,          30,              20,            14,          42,             22]
for i, (h, w) in enumerate(zip(headers2, widths2), start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w
    c = ws2.cell(row=2, column=i, value=h)
    c.font      = Font(name='Arial', bold=True, color=BLANCO, size=10)
    c.fill      = PatternFill('solid', fgColor=AZUL_MED)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border    = borde_thin

ref_data = [
    ('Verano',           'Sep, Oct, Nov, Dic (año ant.) + Ene, Feb',
     'Mar, Abr, May, Jun, Jul, Ago',
     '90-120 días',
     'Producto de temporada estival. Forecast solo en meses activos.',
     'Azul claro'),
    ('Invierno',         'Mar, Abr, May, Jun, Jul, Ago',
     'Ene, Feb, Sep, Oct, Nov, Dic',
     '90-120 días',
     'Producto de temporada invernal. Forecast solo en meses activos.',
     'Verde claro'),
    ('No Estacional',    'Todos los meses',
     'Ninguno',
     '90-120 días',
     'Producto que se vende todo el año de forma uniforme.',
     'Sin color especial'),
    ('Verano/Rotativo',  'Todos los meses (pico en verano)',
     'Ninguno',
     '90-120 días',
     'Se vende todo el año pero con mayor demanda en temporada estival. '
     'El sistema aplica índice estacional al proyectar.',
     'Sin color especial'),
]

for i, row_data in enumerate(ref_data, start=3):
    bg = GRIS if i % 2 == 0 else BLANCO
    for j, val in enumerate(row_data, start=1):
        c = ws2.cell(row=i, column=j, value=val)
        c.font      = Font(name='Arial', size=10)
        c.fill      = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.border    = borde_thin
    ws2.row_dimensions[i].height = 42

# ── Tabla de meses por temporada ─────────────────────────────────────────────
row_m = len(ref_data) + 5
ws2.merge_cells(f'A{row_m}:N{row_m}')
c = ws2.cell(row=row_m, column=1, value='MESES ACTIVOS POR TEMPORADA (✓ = con forecast  |  — = debe ser 0)')
c.font      = Font(name='Arial', bold=True, size=11, color=BLANCO)
c.fill      = PatternFill('solid', fgColor=AZUL_OSC)
c.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[row_m].height = 22

row_m += 1
ws2.cell(row=row_m, column=1, value='Temporada').font = Font(name='Arial', bold=True, size=10)
for i, m in enumerate(MESES, start=2):
    ws2.column_dimensions[get_column_letter(i)].width = 6
    c = ws2.cell(row=row_m, column=i, value=m)
    c.font      = Font(name='Arial', bold=True, color=BLANCO, size=9)
    c.fill      = PatternFill('solid', fgColor=AZUL_MED)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border    = borde_thin
ws2.row_dimensions[row_m].height = 18

meses_activos_map = {
    'Verano':           {1,2,9,10,11,12},
    'Invierno':         {3,4,5,6,7,8},
    'No Estacional':    set(range(1,13)),
    'Verano/Rotativo':  set(range(1,13)),
}

for ti, temp in enumerate(TEMPORADAS_VALIDAS):
    row_m += 1
    bg = GRIS if ti % 2 == 0 else BLANCO
    ws2.cell(row=row_m, column=1, value=temp).font = Font(name='Arial', bold=True, size=10)
    activos = meses_activos_map[temp]
    for mes_n in range(1, 13):
        col = mes_n + 1
        activo = mes_n in activos
        c = ws2.cell(row=row_m, column=col, value='✓' if activo else '—')
        c.font      = Font(name='Arial', size=11, color='375623' if activo else 'AAAAAA', bold=activo)
        c.fill      = PatternFill('solid', fgColor=VERDE_CLAR if activo else bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = borde_thin
    ws2.row_dimensions[row_m].height = 18

ws2.column_dimensions['A'].width = 20

# ══════════════════════════════════════════════════════════════════════════════
# Guardar
# ══════════════════════════════════════════════════════════════════════════════
path = r'C:\Users\rafae\OneDrive\Escritorio\Proyecto Forecast\Plantilla_Forecast_2026.xlsx'
wb.save(path)
print('Guardado en:', path)
