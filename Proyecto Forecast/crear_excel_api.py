from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

AZUL_OSC   = '1F3864'
AZUL_MED   = '2E75B6'
VERDE      = '375623'
VERDE_CLAR = 'E2EFDA'
AMBAR      = 'C55A11'
AMBAR_CLAR = 'FCE4D6'
GRIS       = 'F2F2F2'
BLANCO     = 'FFFFFF'

thin  = Side(style='thin', color='BFBFBF')
borde = Border(left=thin, right=thin, top=thin, bottom=thin)

def hc(ws, row, col, value, bg=AZUL_OSC, fg=BLANCO, size=11, bold=True, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name='Arial', bold=bold, color=fg, size=size)
    c.fill      = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=wrap)
    c.border    = borde
    return c

def dc(ws, row, col, value, bold=False, bg=BLANCO, fg='000000', wrap=False, center=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name='Arial', bold=bold, color=fg, size=10)
    c.fill      = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='center' if center else 'left', vertical='center', wrap_text=wrap)
    c.border    = borde
    return c

# ─── HOJA 1: Especificacion ───────────────────────────────────────────────────
ws1 = wb.active
ws1.title = 'Especificacion API'

ws1.merge_cells('A1:G1')
c = ws1['A1']
c.value = 'ESPECIFICACION API  -  VENTAS 2024 / 2025 / 2026'
c.font  = Font(name='Arial', bold=True, size=14, color=BLANCO)
c.fill  = PatternFill('solid', fgColor=AZUL_OSC)
c.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 30

ws1.merge_cells('A2:G2')
c2 = ws1['A2']
c2.value = 'DCIC - Sistema de Forecast  |  Fecha de solicitud: 10-Jun-2026'
c2.font  = Font(name='Arial', italic=True, size=10, color='595959')
c2.fill  = PatternFill('solid', fgColor='D9E1F2')
c2.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[2].height = 18

ws1.merge_cells('A3:G3')
c3 = ws1['A3']
c3.value = '  OBLIGATORIO = campo imprescindible para forecast y alertas     OPCIONAL = enriquece el analisis pero no bloquea la carga'
c3.font  = Font(name='Arial', size=9, color='404040')
c3.fill  = PatternFill('solid', fgColor='FFF2CC')
c3.alignment = Alignment(horizontal='left', vertical='center')
ws1.row_dimensions[3].height = 16
ws1.row_dimensions[4].height = 6

headers = ['#', 'Nombre Campo API', 'Tipo de Dato', 'Prioridad', 'Formato / Ejemplo', 'Descripcion', 'Observaciones']
col_w   = [5,    28,                 16,             13,          30,                   44,             40]

for i, (h, w) in enumerate(zip(headers, col_w), start=1):
    hc(ws1, 5, i, h, bg=AZUL_MED)
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.row_dimensions[5].height = 22

campos = [
    ('sku',                 'VARCHAR(50)',   'OBLIGATORIO', 'Ej: DCX-001',
     'Codigo unico del producto',
     'Debe coincidir exactamente con el SKU del catalogo DCIC'),

    ('fecha',               'DATE',          'OBLIGATORIO', 'YYYY-MM-DD  Ej: 2025-03-15',
     'Fecha exacta de la transaccion de venta',
     'Granularidad diaria. Sin fecha no se calculan las ultimas 6 semanas de ventas'),

    ('canal',               'VARCHAR(50)',   'OBLIGATORIO', 'ML / Falabella / Web / Directo / Otro',
     'Canal de venta donde se realizo la transaccion',
     'Normalizar valores. El sistema genera alertas separadas por canal'),

    ('cantidad',            'INTEGER',       'OBLIGATORIO', 'Ej: 5  (>= 0)',
     'Unidades vendidas brutas en la transaccion',
     'No descontar devoluciones aqui. Enviar siempre el bruto'),

    ('unidades_devueltas',  'INTEGER',       'OBLIGATORIO', 'Ej: 1   (0 si no hay devolucion)',
     'Unidades devueltas asociadas a esta venta',
     'Venta neta = cantidad - unidades_devueltas. Enviar 0 si no hubo devolucion'),

    ('precio_lista_bruto',  'NUMERIC(12,2)', 'OBLIGATORIO', 'Ej: 29990.00',
     'Precio de lista sin descuento (precio publicado bruto)',
     'Permite identificar ventas con promocion vs precio normal para el forecast'),

    ('valor_unitario_neto', 'NUMERIC(12,2)', 'OBLIGATORIO', 'Ej: 21990.00',
     'Precio efectivo de venta por unidad (despues de descuento)',
     'Ingreso real por unidad. Base para calculo de ingresos del forecast'),

    ('costo_unitario_clp',  'NUMERIC(12,2)', 'OBLIGATORIO', 'Ej: 12500.00',
     'Costo de la unidad en CLP al momento de la venta',
     'Incluir flete e internacion si corresponde'),

    ('margen_clp',          'NUMERIC(12,2)', 'OBLIGATORIO', 'Ej: 9490.00',
     'Margen absoluto en CLP por unidad  (valor_neto - costo)',
     'Si difiere del calculo interno, se usa el valor enviado por la API'),

    ('margen_pct',          'NUMERIC(6,4)',  'OBLIGATORIO', 'Ej: 0.4315  (= 43.15%)',
     'Margen porcentual por unidad en formato decimal',
     'Formato decimal entre -1 y 1. Ej: 0.4315 equivale a 43.15%'),

    ('descripcion_producto','TEXT',          'OPCIONAL',    'Ej: Tabla SUP 10.6 Aqua',
     'Descripcion del producto segun ERP de origen',
     'Solo para validar que el SKU mapeo correctamente al catalogo DCIC'),

    ('categoria_erp',       'VARCHAR(100)',  'OPCIONAL',    'Ej: Tablas SUP',
     'Categoria del producto en el ERP de origen',
     'Para cruzar con catalogo DCIC y detectar inconsistencias de clasificacion'),

    ('marca_erp',           'VARCHAR(100)',  'OPCIONAL',    'Ej: Aquatone',
     'Marca del producto en el ERP de origen',
     'Para validacion de datos entre sistemas'),
]

for i, (nombre, tipo, prioridad, fmt, desc, obs) in enumerate(campos, start=1):
    row   = i + 5
    bg_r  = GRIS if i % 2 == 0 else BLANCO
    bg_p  = VERDE_CLAR if prioridad == 'OBLIGATORIO' else AMBAR_CLAR
    fg_p  = VERDE      if prioridad == 'OBLIGATORIO' else AMBAR

    dc(ws1, row, 1, i,        bold=True, bg=bg_r, center=True)
    dc(ws1, row, 2, nombre,   bold=True, bg=bg_r)
    dc(ws1, row, 3, tipo,     bg=bg_r, center=True)
    dc(ws1, row, 4, prioridad,bold=True, bg=bg_p, fg=fg_p, center=True)
    dc(ws1, row, 5, fmt,      bg=bg_r, wrap=True)
    dc(ws1, row, 6, desc,     bg=bg_r, wrap=True)
    dc(ws1, row, 7, obs,      bg=bg_r, wrap=True)
    ws1.row_dimensions[row].height = 38

ws1.freeze_panes = 'A6'

# ─── HOJA 2: Ejemplo de datos ─────────────────────────────────────────────────
ws2 = wb.create_sheet('Ejemplo Datos')

ws2.merge_cells('A1:M1')
c = ws2['A1']
c.value = 'EJEMPLO DE ESTRUCTURA DE DATOS ESPERADA (valores ficticios)'
c.font  = Font(name='Arial', bold=True, size=13, color=BLANCO)
c.fill  = PatternFill('solid', fgColor=AZUL_OSC)
c.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 26

cols_ej  = ['sku','fecha','canal','cantidad','unidades_devueltas','precio_lista_bruto',
            'valor_unitario_neto','costo_unitario_clp','margen_clp','margen_pct',
            'descripcion_producto','categoria_erp','marca_erp']
anchos_ej = [14,14,14,12,20,20,20,20,14,13,28,18,16]

for i,(h,w) in enumerate(zip(cols_ej, anchos_ej), start=1):
    hc(ws2, 2, i, h, bg=AZUL_MED)
    ws2.column_dimensions[get_column_letter(i)].width = w

ejemplos = [
    ['DCX-001','2025-03-15','ML',         5, 0, 29990.00, 21990.00, 12500.00,  9490.00, 0.4315, 'Tabla SUP 10.6 Aqua', 'Tablas SUP', 'Aquatone'],
    ['DCX-001','2025-03-16','Falabella',  3, 1, 29990.00, 23990.00, 12500.00, 11490.00, 0.4789, 'Tabla SUP 10.6 Aqua', 'Tablas SUP', 'Aquatone'],
    ['DCX-045','2025-06-10','Web',        2, 0, 15990.00, 15990.00,  8200.00,  7790.00, 0.4872, 'Chaleco Neopreno M',  'Neoprenos',  'Rip Curl'],
    ['DCX-112','2024-11-25','ML',        10, 2, 49990.00, 39990.00, 22000.00, 17990.00, 0.4499, 'Kayak Sit-On-Top',    'Kayaks',     'Intex'],
    ['DCX-045','2026-01-08','Directo',    1, 0, 15990.00, 15990.00,  8200.00,  7790.00, 0.4872, 'Chaleco Neopreno M',  'Neoprenos',  'Rip Curl'],
]

for i, fila in enumerate(ejemplos, start=3):
    bg = GRIS if i % 2 == 0 else BLANCO
    for j, val in enumerate(fila, start=1):
        dc(ws2, i, j, val, bg=bg)

ws2.freeze_panes = 'A3'

# ─── HOJA 3: Reglas de validacion ────────────────────────────────────────────
ws3 = wb.create_sheet('Reglas de Validacion')

ws3.merge_cells('A1:E1')
c = ws3['A1']
c.value = 'REGLAS DE VALIDACION Y CONSIDERACIONES'
c.font  = Font(name='Arial', bold=True, size=13, color=BLANCO)
c.fill  = PatternFill('solid', fgColor=AZUL_OSC)
c.alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 26

reg_headers = ['#', 'Campo', 'Regla', 'Error si...', 'Accion esperada']
reg_w       = [5,    20,      40,      36,             36]
for i,(h,w) in enumerate(zip(reg_headers, reg_w), start=1):
    hc(ws3, 2, i, h, bg=AZUL_MED)
    ws3.column_dimensions[get_column_letter(i)].width = w

reglas = [
    ('sku',                 'Debe existir en catalogo DCIC',
     'SKU no reconocido en catalogo',          'Enviar listado de SKUs para homologacion previa'),
    ('fecha',               'Formato YYYY-MM-DD, rango 2024-01-01 a 2026-12-31',
     'Fecha fuera de rango o formato invalido', 'Rechazar registro'),
    ('canal',               'Valores: ML / Falabella / Web / Directo / Otro',
     'Valor no listado',                        'Usar "Otro" si no coincide con ninguno'),
    ('cantidad',            'Entero >= 0',
     'Valor negativo o decimal',                'Rechazar registro'),
    ('unidades_devueltas',  'Entero >= 0 y <= cantidad',
     'Devueltas mayor a cantidad vendida',      'Rechazar registro'),
    ('precio_lista_bruto',  'Decimal positivo, hasta 2 decimales',
     'Valor negativo o cero',                   'Advertencia (no bloquea carga)'),
    ('valor_unitario_neto', 'Decimal positivo, <= precio_lista_bruto',
     'Neto mayor a bruto',                      'Advertencia para revision manual'),
    ('margen_pct',          'Decimal entre -1 y 1  (Ej: 0.43 = 43%)',
     'Valor fuera del rango [-1, 1]',           'Rechazar registro'),
    ('periodo',             'Ventas del 01-Ene-2024 al 31-Dic-2026',
     'Datos fuera del periodo solicitado',      'Filtrar antes de enviar'),
    ('granularidad',        'Una fila por transaccion (NO agrupado por mes)',
     'Datos agrupados por semana o mes',        'Solicitar desagregacion diaria obligatoria'),
]

for i, (campo, regla, error, accion) in enumerate(reglas, start=1):
    row = i + 2
    bg  = GRIS if i % 2 == 0 else BLANCO
    dc(ws3, row, 1, i,      bold=True, bg=bg, center=True)
    dc(ws3, row, 2, campo,  bold=True, bg=bg)
    dc(ws3, row, 3, regla,  bg=bg, wrap=True)
    dc(ws3, row, 4, error,  bg=bg, wrap=True)
    dc(ws3, row, 5, accion, bg=bg, wrap=True)
    ws3.row_dimensions[row].height = 36

ws3.freeze_panes = 'A3'

# ─── Guardar ──────────────────────────────────────────────────────────────────
path = r'C:\Users\rafae\OneDrive\Escritorio\Proyecto Forecast\Solicitud_API_Ventas_DCIC.xlsx'
wb.save(path)
print('Guardado en:', path)
